"""
Leveransplan Comparison Engine
Compares delivery sheets against master leveransplan files.
"""
import re
import os
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ────────────────────────────────────────────────────────────────

DEFAULT_ALLOWLIST = {'.pdf', '.docx', '.xlsx', '.dwg', '.dgn', '.ifc', '.zip', '.rvt', '.jpg', '.xls', '.doc', '.pptx'}

KNOWN_EXTENSIONS = {
    '.pdf', '.xlsx', '.xls', '.dwg', '.doc', '.docx', '.zip', '.rvt',
    '.ifc', '.nwd', '.xml', '.csv', '.txt', '.ppt', '.pptx', '.png',
    '.jpg', '.tif', '.dgn', '.stp', '.step', '.skp', '.bak', '.tmp',
    '.log', '.rar', '.7z', '.jpeg', '.gif', '.bmp', '.svg'
}

EXTENSION_STRINGS = sorted([ext.lstrip('.') for ext in KNOWN_EXTENSIONS], key=len, reverse=True)

REVISION_SUFFIXES = re.compile(
    r'[-_](R\d+|REV\d+|rev\d+|[A-C])$', re.IGNORECASE
)

ILLEGAL_CHARS = set('#*?!%&@"\'<>|')

SKIP_SHEETS = {'Leveransplan', 'Blad1', 'Ändringslogg', 'Ändra inte!'}

FILENAME_HEADER_KEYWORDS = {'filnamn', 'filename', 'file name', 'dokumentnamn', 'number'}


# ── Data classes ─────────────────────────────────────────────────────────────

@dataclass
class FileEntry:
    original_filenames: list  # all format variants from delivery sheet
    cleaned_name: str
    status: str = ''  # FOUND, NOT FOUND, FOUND — WRONG PACKAGE, POSSIBLE MATCH — REVISION
    matched_master_entry: str = ''
    found_in_package: str = ''
    matched_column: str = ''
    flags: list = field(default_factory=list)
    delivery_row_numbers: list = field(default_factory=list)


@dataclass
class MasterEntry:
    value: str
    cleaned: str
    sheet_name: str
    column_letter: str
    row_number: int


# ── Cleaning functions ───────────────────────────────────────────────────────

def clean_text(val):
    """Remove hidden chars, normalize whitespace."""
    if val is None:
        return ''
    s = str(val)
    # Remove BOM, zero-width spaces, non-breaking spaces, carriage returns
    s = s.replace('\ufeff', '').replace('\u200b', '').replace('\u00a0', ' ').replace('\r', '')
    s = s.replace('\n', ' ').replace('\t', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def strip_extension(filename):
    """Strip file extension from a filename. Handles both normal and malformed extensions."""
    # Normal extension
    if '.' in filename:
        base, ext = filename.rsplit('.', 1)
        if '.' + ext.lower() in KNOWN_EXTENSIONS:
            return base
    # Malformed extension (no dot) — check if filename ends with known extension string
    lower = filename.lower()
    for ext_str in EXTENSION_STRINGS:
        if lower.endswith(ext_str) and len(filename) > len(ext_str):
            candidate_base = filename[:-len(ext_str)]
            # Only strip if char before extension is alphanumeric or separator
            if candidate_base and candidate_base[-1] in '-_0123456789':
                return candidate_base
    return filename


def get_extension(filename):
    """Extract extension from filename, or empty string."""
    if '.' in filename:
        ext = '.' + filename.rsplit('.', 1)[1].lower()
        if ext in KNOWN_EXTENSIONS:
            return ext
    return ''


def strip_revision(name):
    """Strip revision suffix for fuzzy matching."""
    return REVISION_SUFFIXES.sub('', name)


# ── Anomaly detection ────────────────────────────────────────────────────────

def detect_anomalies(original_filename, cleaned_name):
    """Detect typos and anomalies in a filename."""
    flags = []
    
    # Missing dot before extension
    lower = original_filename.lower()
    for ext_str in EXTENSION_STRINGS:
        if lower.endswith(ext_str) and not lower.endswith('.' + ext_str):
            if len(original_filename) > len(ext_str):
                flags.append(f'Typo — missing dot before extension ({ext_str})')
                break
    
    # Double extension
    ext_count = 0
    remaining = lower
    for _ in range(3):
        for ext_str in EXTENSION_STRINGS:
            if remaining.endswith('.' + ext_str):
                ext_count += 1
                remaining = remaining[:-(len(ext_str) + 1)]
                break
        else:
            break
    if ext_count >= 2:
        flags.append('Typo — double extension')
    
    # Space inside filename (after edge trimming)
    trimmed = original_filename.strip()
    if ' ' in trimmed:
        flags.append('Typo — space inside filename')
    
    # Illegal characters
    for ch in trimmed:
        if ch in ILLEGAL_CHARS:
            flags.append(f'Anomaly — illegal character: {ch}')
    
    # Filename too short
    if len(cleaned_name) < 8:
        flags.append('Flag — unusually short filename')
    
    # Path separator
    if '/' in trimmed or '\\' in trimmed:
        flags.append('Anomaly — path separator in filename')
    
    return flags


# ── Master file parsing ─────────────────────────────────────────────────────

def detect_filename_column(ws, sheet_name):
    """Auto-detect the column containing filenames using header heuristics."""
    # Scan first 10 rows for header keywords
    for row_idx in range(1, min(11, ws.max_row + 1)):
        for col_idx in range(1, min(ws.max_column + 1, 30)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                cleaned = clean_text(val).lower()
                for keyword in FILENAME_HEADER_KEYWORDS:
                    if keyword in cleaned and 'format' not in cleaned:
                        return col_idx, get_column_letter(col_idx), row_idx
    # Fallback: Column E
    return 5, 'E', None


def parse_master_file(filepath):
    """Parse the master leveransplan file and extract all filenames."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    master_entries = []  # list of MasterEntry
    primary_col_entries = defaultdict(list)  # sheet -> list of MasterEntry (primary col only)
    all_col_entries = []  # all entries from all columns
    sheet_info = {}
    
    filename_pattern = re.compile(r'^E\d{2}[- ]', re.IGNORECASE)
    
    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in SKIP_SHEETS:
            continue
        
        ws = wb[sheet_name]
        primary_col_idx, primary_col_letter, header_row = detect_filename_column(ws, sheet_name)
        
        sheet_info[sheet_name] = {
            'primary_col': primary_col_letter,
            'primary_col_idx': primary_col_idx,
            'header_row': header_row,
            'total_rows': ws.max_row,
        }
        
        data_start = (header_row + 1) if header_row else 1
        
        for row_idx in range(data_start, ws.max_row + 1):
            for col_idx in range(1, min(ws.max_column + 1, 30)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if not val:
                    continue
                cleaned_val = clean_text(val)
                if not cleaned_val:
                    continue
                
                # Check if it looks like a document filename
                if not filename_pattern.match(cleaned_val):
                    continue
                
                # Strip extension if present in master (flag as anomaly)
                has_ext = get_extension(cleaned_val) != ''
                cleaned_for_match = strip_extension(cleaned_val) if has_ext else cleaned_val
                cleaned_for_match = clean_text(cleaned_for_match)
                
                col_letter = get_column_letter(col_idx)
                entry = MasterEntry(
                    value=cleaned_val,
                    cleaned=cleaned_for_match.lower(),
                    sheet_name=sheet_name,
                    column_letter=col_letter,
                    row_number=row_idx
                )
                
                all_col_entries.append(entry)
                if col_idx == primary_col_idx:
                    primary_col_entries[sheet_name].append(entry)
    
    wb.close()
    return primary_col_entries, all_col_entries, sheet_info


def parse_delivery_sheet(filepath, allowlist=None):
    """Parse the delivery sheet and extract filenames from Column B."""
    if allowlist is None:
        allowlist = DEFAULT_ALLOWLIST
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Find the main data sheet
    ws = None
    for name in wb.sheetnames:
        if 'delivery' in name.lower() or 'leverans' in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    
    # Detect header row (look for "NUMBER" in column B)
    header_row = 6
    for row_idx in range(1, min(15, ws.max_row + 1)):
        val = ws.cell(row=row_idx, column=2).value
        if val and 'number' in str(val).lower():
            header_row = row_idx
            break
    
    raw_entries = []  # (row_number, original_filename)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=2).value
        if not val:
            continue
        original = clean_text(val)
        if not original:
            continue
        # Skip the package header row (same as package number)
        if original.startswith('E03-CD-') and 'PCG' in original and not get_extension(original):
            # Check if this is likely a header vs a real filename
            title_val = ws.cell(row=row_idx, column=3).value
            if title_val and 'MEP' in str(title_val).upper():
                continue

        raw_entries.append((row_idx, original))

    raw_row_count = len(raw_entries)

    # Apply allowlist filter
    filtered_entries = []
    excluded = []
    excluded_by_ext = defaultdict(int)
    for row_num, original in raw_entries:
        ext = get_extension(original)
        if ext and ext not in allowlist:
            excluded.append((row_num, original, ext))
            excluded_by_ext[ext] += 1
            continue
        filtered_entries.append((row_num, original))

    # Clean and deduplicate (collapse multi-format groups)
    groups = defaultdict(list)  # cleaned_name -> [(row_num, original)]
    for row_num, original in filtered_entries:
        cleaned = strip_extension(original)
        cleaned = clean_text(cleaned)
        groups[cleaned].append((row_num, original))

    multi_format_duplicates = len(filtered_entries) - len(groups)
    unique_files_for_comparison = len(groups)
    
    # Build FileEntry objects
    file_entries = []
    seen_exact = set()  # track true duplicates (same filename+extension)
    
    for cleaned_name, variants in groups.items():
        entry = FileEntry(
            original_filenames=[orig for _, orig in variants],
            cleaned_name=cleaned_name,
            delivery_row_numbers=[rn for rn, _ in variants],
        )
        
        # Detect true duplicates (exact same filename+extension appearing multiple times)
        exact_names = [orig for _, orig in variants]
        seen_in_group = set()
        for orig in exact_names:
            if orig in seen_in_group:
                first_row = next(rn for rn, o in variants if o == orig)
                entry.flags.append(f'Duplicate — identical entry on row {first_row}')
            seen_in_group.add(orig)
        
        # Detect anomalies on each variant
        for orig in entry.original_filenames:
            anomalies = detect_anomalies(orig, cleaned_name)
            for a in anomalies:
                if a not in entry.flags:
                    entry.flags.append(a)
        
        file_entries.append(entry)
    
    pipeline_stats = {
        'raw_row_count': raw_row_count,
        'excluded_by_allowlist': len(excluded),
        'excluded_by_ext': dict(excluded_by_ext),
        'multi_format_duplicates': multi_format_duplicates,
        'unique_files_for_comparison': unique_files_for_comparison,
    }

    wb.close()

    return file_entries, excluded, header_row, ws.title, pipeline_stats


# ── Comparison engine ────────────────────────────────────────────────────────

def determine_expected_package(file_entries, sheet_info):
    """Determine which master sheet/package the delivery belongs to based on filename prefixes."""
    # Count prefix patterns like E03-52, E03-55, E03-63 etc.
    prefix_counts = defaultdict(int)
    for entry in file_entries:
        match = re.match(r'(E\d{2}-\d{2,})', entry.cleaned_name, re.IGNORECASE)
        if match:
            prefix_counts[match.group(1).upper()] += 1
    
    # Map prefixes to sheets based on what filenames exist on each sheet
    return prefix_counts


def run_comparison(master_path, delivery_path, allowlist=None):
    """Run the full comparison and return results."""
    
    # Parse both files
    primary_entries, all_entries, sheet_info = parse_master_file(master_path)
    file_entries, excluded, header_row, delivery_sheet_name, pipeline_stats = parse_delivery_sheet(delivery_path, allowlist)
    
    if not file_entries:
        return None, "No valid filenames found in delivery sheet Column B."
    
    # Build lookup indexes
    # Primary column: cleaned_name -> list of MasterEntry
    primary_index = defaultdict(list)
    for sheet_name, entries in primary_entries.items():
        for e in entries:
            primary_index[e.cleaned].append(e)
    
    # All columns: cleaned_name -> list of MasterEntry
    all_index = defaultdict(list)
    for e in all_entries:
        all_index[e.cleaned].append(e)
    
    # Revision-stripped indexes
    primary_rev_index = defaultdict(list)
    for sheet_name, entries in primary_entries.items():
        for e in entries:
            rev_stripped = strip_revision(e.cleaned)
            primary_rev_index[rev_stripped].append(e)
    
    all_rev_index = defaultdict(list)
    for e in all_entries:
        rev_stripped = strip_revision(e.cleaned)
        all_rev_index[rev_stripped].append(e)
    
    # Run matching for each delivery file
    for entry in file_entries:
        cleaned_lower = entry.cleaned_name.lower()
        
        # Rule 1: Exact match in primary column
        if cleaned_lower in primary_index:
            matches = primary_index[cleaned_lower]
            entry.status = 'FOUND'
            best = matches[0]
            entry.matched_master_entry = best.value
            entry.found_in_package = best.sheet_name
            entry.matched_column = best.column_letter
            continue
        
        # Rule 2: Exact match in any column
        if cleaned_lower in all_index:
            matches = all_index[cleaned_lower]
            entry.status = 'FOUND'
            best = matches[0]
            entry.matched_master_entry = best.value
            entry.found_in_package = best.sheet_name
            entry.matched_column = best.column_letter
            if best.column_letter != 'E':
                entry.flags.append(f'Note — matched outside Column E (Col {best.column_letter}, Sheet {best.sheet_name})')
            continue
        
        # Rule 3: Check if found but wrong package (this applies if we know expected package)
        # For now, any match found is valid since we don't know expected package per-file
        
        # Rule 4: Fuzzy revision match
        rev_stripped = strip_revision(cleaned_lower)
        if rev_stripped != cleaned_lower:
            if rev_stripped in primary_rev_index:
                matches = primary_rev_index[rev_stripped]
                entry.status = 'POSSIBLE MATCH — REVISION'
                best = matches[0]
                entry.matched_master_entry = best.value
                entry.found_in_package = best.sheet_name
                entry.matched_column = best.column_letter
                continue
            if rev_stripped in all_rev_index:
                matches = all_rev_index[rev_stripped]
                entry.status = 'POSSIBLE MATCH — REVISION'
                best = matches[0]
                entry.matched_master_entry = best.value
                entry.found_in_package = best.sheet_name
                entry.matched_column = best.column_letter
                continue
        
        # Also try: master entries have revision but delivery doesn't
        found_rev = False
        for master_cleaned, master_list in primary_rev_index.items():
            if master_cleaned == cleaned_lower:
                entry.status = 'POSSIBLE MATCH — REVISION'
                best = master_list[0]
                entry.matched_master_entry = best.value
                entry.found_in_package = best.sheet_name
                entry.matched_column = best.column_letter
                found_rev = True
                break
        if found_rev:
            continue
        
        # Rule 5: Not found
        entry.status = 'NOT FOUND'
    
    # Compute statistics
    total_unique = len(file_entries)
    found_count = sum(1 for e in file_entries if e.status == 'FOUND')
    not_found_count = sum(1 for e in file_entries if e.status == 'NOT FOUND')
    wrong_package_count = sum(1 for e in file_entries if e.status == 'FOUND — WRONG PACKAGE')
    revision_count = sum(1 for e in file_entries if e.status == 'POSSIBLE MATCH — REVISION')
    
    # Count master primary column entries
    master_primary_count = sum(len(entries) for entries in primary_entries.values())
    
    flagged_count = sum(1 for e in file_entries if e.flags)
    duplicate_count = sum(1 for e in file_entries if any('Duplicate' in f for f in e.flags))
    
    match_rate = (found_count / total_unique * 100) if total_unique > 0 else 0
    
    stats = {
        'total_unique': total_unique,
        'found': found_count,
        'not_found': not_found_count,
        'wrong_package': wrong_package_count,
        'revision_match': revision_count,
        'match_rate': match_rate,
        'flagged': flagged_count,
        'duplicates': duplicate_count,
        'master_primary_count': master_primary_count,
        'excluded_count': len(excluded),
        'excluded_files': excluded,
        # Pipeline funnel counts
        'raw_row_count': pipeline_stats['raw_row_count'],
        'excluded_by_allowlist': pipeline_stats['excluded_by_allowlist'],
        'excluded_by_ext': pipeline_stats['excluded_by_ext'],
        'multi_format_duplicates': pipeline_stats['multi_format_duplicates'],
        'unique_files_for_comparison': pipeline_stats['unique_files_for_comparison'],
    }
    
    return {
        'file_entries': file_entries,
        'stats': stats,
        'sheet_info': sheet_info,
        'delivery_sheet_name': delivery_sheet_name,
    }, None


# ── Excel report generation ─────────────────────────────────────────────────

# Colors
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
ORANGE_FILL = PatternFill('solid', fgColor='FFCC99')
YELLOW_FILL = PatternFill('solid', fgColor='FFEB9C')
LIGHT_YELLOW_FILL = PatternFill('solid', fgColor='FFFFCC')
LIGHT_RED_FILL = PatternFill('solid', fgColor='FFE0E0')
LIGHT_GREEN_FILL = PatternFill('solid', fgColor='E2F0D9')
SECTION_FILL = PatternFill('solid', fgColor='2D3F5F')
HEADER_FILL = PatternFill('solid', fgColor='1B2A4A')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(bold=True, name='Arial', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_body_cell(ws, row, col, font=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font or BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical='top', wrap_text=True)
    return cell


def generate_report(results, output_path):
    """Generate the comparison report Excel file."""
    wb = openpyxl.Workbook()
    entries = results['file_entries']
    stats = results['stats']
    
    # ── TAB 1: Missing Files ────────────────────────────────────────────
    ws_missing = wb.active
    ws_missing.title = 'Missing Files'
    
    missing = [e for e in entries if e.status != 'FOUND']
    
    headers = ['#', 'Original Filename(s)', 'Cleaned Filename', 'Status', 'Closest Match in Master', 'Package (if found)', 'Flags']
    for col, h in enumerate(headers, 1):
        ws_missing.cell(row=1, column=col, value=h)
    style_header_row(ws_missing, 1, len(headers))
    
    # Sort: NOT FOUND first, then WRONG PACKAGE, then REVISION
    status_order = {'NOT FOUND': 0, 'FOUND — WRONG PACKAGE': 1, 'POSSIBLE MATCH — REVISION': 2}
    missing.sort(key=lambda e: status_order.get(e.status, 99))
    
    for idx, entry in enumerate(missing, 1):
        row = idx + 1
        style_body_cell(ws_missing, row, 1).value = idx
        style_body_cell(ws_missing, row, 2).value = ', '.join(entry.original_filenames)
        style_body_cell(ws_missing, row, 3).value = entry.cleaned_name
        
        status_cell = style_body_cell(ws_missing, row, 4)
        status_cell.value = entry.status
        if entry.status == 'NOT FOUND':
            status_cell.fill = RED_FILL
        elif entry.status == 'FOUND — WRONG PACKAGE':
            status_cell.fill = ORANGE_FILL
        elif entry.status == 'POSSIBLE MATCH — REVISION':
            status_cell.fill = YELLOW_FILL
        
        style_body_cell(ws_missing, row, 5).value = entry.matched_master_entry or ''
        style_body_cell(ws_missing, row, 6).value = entry.found_in_package or ''
        style_body_cell(ws_missing, row, 7).value = '; '.join(entry.flags) if entry.flags else ''
    
    if not missing:
        ws_missing.cell(row=2, column=1, value='No missing files — all deliveries matched!').font = Font(bold=True, color='006100', name='Arial', size=11)
    
    ws_missing.freeze_panes = 'A2'
    for col in range(1, len(headers) + 1):
        ws_missing.column_dimensions[get_column_letter(col)].width = [5, 45, 40, 28, 40, 30, 45][col - 1]
    
    # ── TAB 2: Full Detail ──────────────────────────────────────────────
    ws_detail = wb.create_sheet('Full Detail')
    
    detail_headers = ['Original Filename(s)', 'Cleaned Filename', 'Matched Master Entry', 'Status', 'Found In Package', 'Flags & Notes']
    for col, h in enumerate(detail_headers, 1):
        ws_detail.cell(row=1, column=col, value=h)
    style_header_row(ws_detail, 1, len(detail_headers))
    
    # Sort: NOT FOUND first, then WRONG PACKAGE, REVISION, FOUND last
    all_status_order = {'NOT FOUND': 0, 'FOUND — WRONG PACKAGE': 1, 'POSSIBLE MATCH — REVISION': 2, 'FOUND': 3}
    sorted_entries = sorted(entries, key=lambda e: all_status_order.get(e.status, 99))
    
    for idx, entry in enumerate(sorted_entries, 1):
        row = idx + 1
        style_body_cell(ws_detail, row, 1).value = ', '.join(entry.original_filenames)
        style_body_cell(ws_detail, row, 2).value = entry.cleaned_name
        style_body_cell(ws_detail, row, 3).value = entry.matched_master_entry or ''
        
        status_cell = style_body_cell(ws_detail, row, 4)
        status_cell.value = entry.status
        if entry.status == 'FOUND':
            status_cell.fill = GREEN_FILL
        elif entry.status == 'NOT FOUND':
            status_cell.fill = RED_FILL
        elif entry.status == 'FOUND — WRONG PACKAGE':
            status_cell.fill = ORANGE_FILL
        elif entry.status == 'POSSIBLE MATCH — REVISION':
            status_cell.fill = YELLOW_FILL
        
        # Highlight mismatches between cleaned delivery and master entry
        if entry.matched_master_entry:
            if entry.cleaned_name.lower() != entry.matched_master_entry.lower():
                ws_detail.cell(row=row, column=2).fill = LIGHT_YELLOW_FILL
                ws_detail.cell(row=row, column=3).fill = LIGHT_YELLOW_FILL
        
        style_body_cell(ws_detail, row, 5).value = entry.found_in_package or ''
        style_body_cell(ws_detail, row, 6).value = '; '.join(entry.flags) if entry.flags else ''
    
    ws_detail.freeze_panes = 'A2'
    for col in range(1, len(detail_headers) + 1):
        ws_detail.column_dimensions[get_column_letter(col)].width = [45, 40, 40, 28, 30, 45][col - 1]
    
    # ── TAB 3: Summary / Overview ───────────────────────────────────────
    ws_summary = wb.create_sheet('Summary')

    sum_headers = ['Metric', 'Count', 'Master File', 'Notes / Details']
    for col, h in enumerate(sum_headers, 1):
        ws_summary.cell(row=1, column=col, value=h)
    style_header_row(ws_summary, 1, len(sum_headers))

    # ── Section 1: Processing Pipeline ──────────────────────────────────

    # Section header row
    sec1_row = 2
    ws_summary.cell(row=sec1_row, column=1, value='PROCESSING PIPELINE')
    ws_summary.merge_cells(start_row=sec1_row, start_column=1, end_row=sec1_row, end_column=4)
    sec1_cell = ws_summary.cell(row=sec1_row, column=1)
    sec1_cell.fill = SECTION_FILL
    sec1_cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    sec1_cell.alignment = Alignment(horizontal='left', vertical='center')
    sec1_cell.border = THIN_BORDER

    # Build excluded-by-extension detail string
    exc_by_ext = stats.get('excluded_by_ext', {})
    if exc_by_ext:
        ext_detail = ', '.join(f"{ext}: {cnt}" for ext, cnt in sorted(exc_by_ext.items()))
    else:
        ext_detail = '—'

    pipeline_rows = [
        (
            'Raw entries in Delivery Sheet',
            stats['raw_row_count'],
            '—',
            'Total non-empty rows in Column B',
            None,
        ),
        (
            'Removed — extension not in allowlist',
            f"-{stats['excluded_by_allowlist']}",
            '—',
            ext_detail,
            LIGHT_RED_FILL,
        ),
        (
            'Removed — multi-format deduplication',
            f"-{stats['multi_format_duplicates']}",
            '—',
            'Same base name, different extensions collapsed to 1',
            LIGHT_RED_FILL,
        ),
        (
            'Files entering comparison',
            stats['unique_files_for_comparison'],
            '—',
            'Actual comparison input',
            LIGHT_GREEN_FILL,
        ),
    ]

    THICK_BOTTOM = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='medium', color='4472C4'),
    )

    for i, (metric, count_val, master_val, notes, fill) in enumerate(pipeline_rows):
        row = sec1_row + 1 + i
        is_final = (i == len(pipeline_rows) - 1)
        font = BOLD_FONT if is_final else BODY_FONT
        border = THICK_BOTTOM if is_final else THIN_BORDER

        for col in range(1, 5):
            cell = ws_summary.cell(row=row, column=col)
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if fill:
                cell.fill = fill

        ws_summary.cell(row=row, column=1).value = metric
        ws_summary.cell(row=row, column=2).value = count_val
        ws_summary.cell(row=row, column=3).value = master_val
        ws_summary.cell(row=row, column=4).value = notes

    # Blank spacer row
    spacer_row = sec1_row + len(pipeline_rows) + 1
    for col in range(1, 5):
        ws_summary.cell(row=spacer_row, column=col).border = Border()

    # ── Section 2: Comparison Results ───────────────────────────────────

    sec2_row = spacer_row + 1
    ws_summary.cell(row=sec2_row, column=1, value='COMPARISON RESULTS')
    ws_summary.merge_cells(start_row=sec2_row, start_column=1, end_row=sec2_row, end_column=4)
    sec2_cell = ws_summary.cell(row=sec2_row, column=1)
    sec2_cell.fill = SECTION_FILL
    sec2_cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    sec2_cell.alignment = Alignment(horizontal='left', vertical='center')
    sec2_cell.border = THIN_BORDER

    unique = stats['unique_files_for_comparison']
    delta = unique - stats['master_primary_count']
    delta_str = f"{'+' if delta > 0 else ''}{delta}"

    comparison_rows = [
        ('Total files compared', str(unique), str(stats['master_primary_count']), delta_str),
        ('Files confirmed present (FOUND)', str(stats['found']), '—', '—'),
        ('Files missing from Master', str(stats['not_found']), '—', 'Action required' if stats['not_found'] > 0 else '—'),
        ('Files under wrong package', str(stats['wrong_package']), '—', 'Review needed' if stats['wrong_package'] > 0 else '—'),
        ('Possible revision matches', str(stats['revision_match']), '—', 'Manual check' if stats['revision_match'] > 0 else '—'),
        ('Match Rate', f"{stats['match_rate']:.1f}%", '—', ''),
        ('Typos / Anomalies detected', str(stats['flagged']), '—', 'See Detail tab' if stats['flagged'] > 0 else '—'),
        ('Duplicates detected', str(stats['duplicates']), '—', '—'),
    ]

    for idx, (metric, delivery_val, master_val, delta_note) in enumerate(comparison_rows):
        row = sec2_row + 1 + idx
        style_body_cell(ws_summary, row, 1, font=BOLD_FONT).value = metric
        style_body_cell(ws_summary, row, 2).value = delivery_val
        style_body_cell(ws_summary, row, 3).value = master_val
        style_body_cell(ws_summary, row, 4).value = delta_note

        # Color coding
        if metric == 'Files missing from Master' and stats['not_found'] > 0:
            ws_summary.cell(row=row, column=2).fill = RED_FILL
            ws_summary.cell(row=row, column=4).fill = RED_FILL
        elif metric == 'Files under wrong package' and stats['wrong_package'] > 0:
            ws_summary.cell(row=row, column=2).fill = ORANGE_FILL
            ws_summary.cell(row=row, column=4).fill = ORANGE_FILL
        elif metric == 'Match Rate':
            rate = stats['match_rate']
            if rate >= 99:
                ws_summary.cell(row=row, column=2).fill = GREEN_FILL
            elif rate >= 95:
                ws_summary.cell(row=row, column=2).fill = YELLOW_FILL
            else:
                ws_summary.cell(row=row, column=2).fill = RED_FILL
        elif metric in ('Typos / Anomalies detected', 'Duplicates detected'):
            val = stats['flagged'] if 'Typo' in metric else stats['duplicates']
            if val > 0:
                ws_summary.cell(row=row, column=2).fill = YELLOW_FILL

    ws_summary.freeze_panes = 'A2'
    for col in range(1, 5):
        ws_summary.column_dimensions[get_column_letter(col)].width = [40, 20, 20, 45][col - 1]
    
    # Move Summary to first position
    wb.move_sheet('Summary', offset=-2)
    
    wb.save(output_path)
    return output_path


# ── Main API function ────────────────────────────────────────────────────────

def compare_and_report(master_path, delivery_path, output_path, allowlist=None):
    """Full pipeline: parse, compare, generate report."""
    results, error = run_comparison(master_path, delivery_path, allowlist)
    if error:
        return None, error
    
    report_path = generate_report(results, output_path)
    
    # Build text summary
    entries = results['file_entries']
    stats = results['stats']
    
    lines = []
    lines.append(f"COMPARISON SUMMARY")
    lines.append(f"{'=' * 50}")

    # Pipeline funnel
    lines.append(f"PROCESSING PIPELINE")
    lines.append(f"  Raw entries in delivery sheet:     {stats['raw_row_count']}")
    exc_detail = ''
    if stats.get('excluded_by_ext'):
        exc_detail = '  (' + ', '.join(f"{ext}: {cnt}" for ext, cnt in sorted(stats['excluded_by_ext'].items())) + ')'
    lines.append(f"  Removed — not in allowlist:       -{stats['excluded_by_allowlist']}{exc_detail}")
    lines.append(f"  Removed — multi-format duplicates: -{stats['multi_format_duplicates']}")
    lines.append(f"  → Files entering comparison:       {stats['unique_files_for_comparison']}")
    lines.append(f"")

    lines.append(f"COMPARISON RESULTS")
    lines.append(f"  Match rate: {stats['match_rate']:.1f}%")
    lines.append(f"")
    
    not_found = [e for e in entries if e.status == 'NOT FOUND']
    lines.append(f"NOT FOUND ({len(not_found)}):")
    if not_found:
        for e in not_found:
            flag_str = f" [{'; '.join(e.flags)}]" if e.flags else ""
            lines.append(f"  • {e.cleaned_name}{flag_str}")
    else:
        lines.append("  None detected.")
    lines.append("")
    
    wrong_pkg = [e for e in entries if e.status == 'FOUND — WRONG PACKAGE']
    lines.append(f"FOUND — WRONG PACKAGE ({len(wrong_pkg)}):")
    if wrong_pkg:
        for e in wrong_pkg:
            lines.append(f"  • {e.cleaned_name} → found in: {e.found_in_package}")
    else:
        lines.append("  None detected.")
    lines.append("")
    
    rev_match = [e for e in entries if e.status == 'POSSIBLE MATCH — REVISION']
    lines.append(f"POSSIBLE MATCH — REVISION ({len(rev_match)}):")
    if rev_match:
        for e in rev_match:
            lines.append(f"  • {e.cleaned_name} ↔ {e.matched_master_entry}")
    else:
        lines.append("  None detected.")
    lines.append("")
    
    flagged = [e for e in entries if e.flags]
    lines.append(f"TYPOS / ANOMALIES ({len(flagged)}):")
    if flagged:
        for e in flagged:
            lines.append(f"  • {e.cleaned_name}: {'; '.join(e.flags)}")
    else:
        lines.append("  None detected.")
    
    summary_text = '\n'.join(lines)
    
    return {
        'report_path': report_path,
        'summary_text': summary_text,
        'stats': stats,
    }, None


# ── Multi-delivery report ────────────────────────────────────────────────────

def _safe_tab_name(filename_stem, suffix, max_stem=28):
    """Build an Excel tab name within the 31-char limit."""
    stem = filename_stem[:max_stem]
    return f"{stem} {suffix}"


def generate_combined_report(results_list, output_path):
    """Generate a single Excel report for multiple delivery sheets.

    results_list: list of (delivery_filename, run_comparison result dict)
    """
    wb = openpyxl.Workbook()

    # ── Aggregate Summary tab ───────────────────────────────────────────
    ws_agg = wb.active
    ws_agg.title = 'Aggregate Summary'

    agg_headers = ['Delivery Sheet', 'Raw Entries', 'Compared', 'Found', 'Not Found', 'Match Rate', 'Revisions', 'Anomalies']
    for col, h in enumerate(agg_headers, 1):
        ws_agg.cell(row=1, column=col, value=h)
    style_header_row(ws_agg, 1, len(agg_headers))

    totals = {'raw': 0, 'compared': 0, 'found': 0, 'not_found': 0, 'revision': 0, 'anomalies': 0}

    for data_row_idx, (delivery_filename, result) in enumerate(results_list, 1):
        stats = result['stats']
        row = data_row_idx + 1
        rate = stats['match_rate']

        style_body_cell(ws_agg, row, 1).value = delivery_filename
        style_body_cell(ws_agg, row, 2).value = stats['raw_row_count']
        style_body_cell(ws_agg, row, 3).value = stats['unique_files_for_comparison']
        style_body_cell(ws_agg, row, 4).value = stats['found']
        nf_cell = style_body_cell(ws_agg, row, 5)
        nf_cell.value = stats['not_found']
        if stats['not_found'] > 0:
            nf_cell.fill = RED_FILL

        rate_cell = style_body_cell(ws_agg, row, 6)
        rate_cell.value = f"{rate:.1f}%"
        if rate >= 99:
            rate_cell.fill = GREEN_FILL
        elif rate >= 95:
            rate_cell.fill = YELLOW_FILL
        else:
            rate_cell.fill = RED_FILL

        rev_cell = style_body_cell(ws_agg, row, 7)
        rev_cell.value = stats['revision_match']
        if stats['revision_match'] > 0:
            rev_cell.fill = YELLOW_FILL

        anom_cell = style_body_cell(ws_agg, row, 8)
        anom_cell.value = stats['flagged']
        if stats['flagged'] > 0:
            anom_cell.fill = YELLOW_FILL

        totals['raw'] += stats['raw_row_count']
        totals['compared'] += stats['unique_files_for_comparison']
        totals['found'] += stats['found']
        totals['not_found'] += stats['not_found']
        totals['revision'] += stats['revision_match']
        totals['anomalies'] += stats['flagged']

    # Totals row
    totals_row = len(results_list) + 2
    total_compared = totals['compared']
    avg_rate = (totals['found'] / total_compared * 100) if total_compared > 0 else 0
    totals_data = ['TOTAL', totals['raw'], totals['compared'], totals['found'], totals['not_found'],
                   f"{avg_rate:.1f}%", totals['revision'], totals['anomalies']]
    for col, val in enumerate(totals_data, 1):
        cell = ws_agg.cell(row=totals_row, column=col, value=val)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='top')

    ws_agg.freeze_panes = 'A2'
    agg_col_widths = [45, 14, 14, 10, 14, 14, 14, 14]
    for col, width in enumerate(agg_col_widths, 1):
        ws_agg.column_dimensions[get_column_letter(col)].width = width

    # ── Per-delivery tabs ───────────────────────────────────────────────
    missing_headers = ['#', 'Original Filename(s)', 'Cleaned Filename', 'Status', 'Closest Match in Master', 'Package (if found)', 'Flags']
    detail_headers = ['Original Filename(s)', 'Cleaned Filename', 'Matched Master Entry', 'Status', 'Found In Package', 'Flags & Notes']

    status_order = {'NOT FOUND': 0, 'FOUND — WRONG PACKAGE': 1, 'POSSIBLE MATCH — REVISION': 2}
    all_status_order = {'NOT FOUND': 0, 'FOUND — WRONG PACKAGE': 1, 'POSSIBLE MATCH — REVISION': 2, 'FOUND': 3}

    for delivery_filename, result in results_list:
        entries = result['file_entries']
        stem = Path(delivery_filename).stem

        # Missing Files tab
        ws_m = wb.create_sheet(_safe_tab_name(stem, '— Missing'))
        for col, h in enumerate(missing_headers, 1):
            ws_m.cell(row=1, column=col, value=h)
        style_header_row(ws_m, 1, len(missing_headers))

        missing = [e for e in entries if e.status != 'FOUND']
        missing.sort(key=lambda e: status_order.get(e.status, 99))

        if not missing:
            ws_m.cell(row=2, column=1, value='No missing files — all deliveries matched!').font = Font(bold=True, color='006100', name='Arial', size=11)
        else:
            for idx, entry in enumerate(missing, 1):
                row = idx + 1
                style_body_cell(ws_m, row, 1).value = idx
                style_body_cell(ws_m, row, 2).value = ', '.join(entry.original_filenames)
                style_body_cell(ws_m, row, 3).value = entry.cleaned_name
                status_cell = style_body_cell(ws_m, row, 4)
                status_cell.value = entry.status
                if entry.status == 'NOT FOUND':
                    status_cell.fill = RED_FILL
                elif entry.status == 'FOUND — WRONG PACKAGE':
                    status_cell.fill = ORANGE_FILL
                elif entry.status == 'POSSIBLE MATCH — REVISION':
                    status_cell.fill = YELLOW_FILL
                style_body_cell(ws_m, row, 5).value = entry.matched_master_entry or ''
                style_body_cell(ws_m, row, 6).value = entry.found_in_package or ''
                style_body_cell(ws_m, row, 7).value = '; '.join(entry.flags) if entry.flags else ''

        ws_m.freeze_panes = 'A2'
        for col, width in enumerate([5, 45, 40, 28, 40, 30, 45], 1):
            ws_m.column_dimensions[get_column_letter(col)].width = width

        # Full Detail tab
        ws_d = wb.create_sheet(_safe_tab_name(stem, '— Detail'))
        for col, h in enumerate(detail_headers, 1):
            ws_d.cell(row=1, column=col, value=h)
        style_header_row(ws_d, 1, len(detail_headers))

        sorted_entries = sorted(entries, key=lambda e: all_status_order.get(e.status, 99))
        for idx, entry in enumerate(sorted_entries, 1):
            row = idx + 1
            style_body_cell(ws_d, row, 1).value = ', '.join(entry.original_filenames)
            style_body_cell(ws_d, row, 2).value = entry.cleaned_name
            style_body_cell(ws_d, row, 3).value = entry.matched_master_entry or ''
            status_cell = style_body_cell(ws_d, row, 4)
            status_cell.value = entry.status
            if entry.status == 'FOUND':
                status_cell.fill = GREEN_FILL
            elif entry.status == 'NOT FOUND':
                status_cell.fill = RED_FILL
            elif entry.status == 'FOUND — WRONG PACKAGE':
                status_cell.fill = ORANGE_FILL
            elif entry.status == 'POSSIBLE MATCH — REVISION':
                status_cell.fill = YELLOW_FILL
            if entry.matched_master_entry and entry.cleaned_name.lower() != entry.matched_master_entry.lower():
                ws_d.cell(row=row, column=2).fill = LIGHT_YELLOW_FILL
                ws_d.cell(row=row, column=3).fill = LIGHT_YELLOW_FILL
            style_body_cell(ws_d, row, 5).value = entry.found_in_package or ''
            style_body_cell(ws_d, row, 6).value = '; '.join(entry.flags) if entry.flags else ''

        ws_d.freeze_panes = 'A2'
        for col, width in enumerate([45, 40, 40, 28, 30, 45], 1):
            ws_d.column_dimensions[get_column_letter(col)].width = width

    wb.save(output_path)
    return output_path


def compare_and_report_multiple(master_path, delivery_paths, output_path, allowlist=None):
    """Full pipeline for multiple delivery sheets: parse, compare each, generate combined report."""
    results_list = []

    for delivery_path in delivery_paths:
        result, error = run_comparison(master_path, delivery_path, allowlist)
        if error:
            return None, f"Error processing {delivery_path}: {error}"
        delivery_filename = Path(delivery_path).name
        results_list.append((delivery_filename, result))

    report_path = generate_combined_report(results_list, output_path)

    # Build combined summary text
    lines = ['MULTI-DELIVERY COMPARISON SUMMARY', '=' * 50, '']
    total_compared = 0
    total_found = 0

    for delivery_filename, result in results_list:
        stats = result['stats']
        total_compared += stats['unique_files_for_comparison']
        total_found += stats['found']
        lines.append(f"── {delivery_filename}")
        lines.append(f"   Raw: {stats['raw_row_count']}  Compared: {stats['unique_files_for_comparison']}  Found: {stats['found']}  Not Found: {stats['not_found']}  Match Rate: {stats['match_rate']:.1f}%")
        if stats['not_found'] > 0:
            not_found = [e for e in result['file_entries'] if e.status == 'NOT FOUND']
            for e in not_found:
                lines.append(f"   • NOT FOUND: {e.cleaned_name}")
        lines.append('')

    overall_rate = (total_found / total_compared * 100) if total_compared > 0 else 0
    lines.append(f"OVERALL: {total_found}/{total_compared} files matched ({overall_rate:.1f}%)")

    summary_text = '\n'.join(lines)
    stats_list = [{'delivery': fn, 'stats': r['stats']} for fn, r in results_list]

    return {
        'report_path': report_path,
        'summary_text': summary_text,
        'stats_list': stats_list,
    }, None
